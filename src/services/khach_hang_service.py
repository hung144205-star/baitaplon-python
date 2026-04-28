#!/usr/bin/env python3
"""
Khách hàng Service - Business logic cho Quản lý Khách hàng
"""
from typing import List, Optional, Dict, Any
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import or_

from src.models import KhachHang, HopDong, ThanhToan, LoaiKhachEnum, TrangThaiKHEnum
from src.database import get_session, session_scope
from src.utils.validators import validate_email, validate_phone, validate_required


class KhachHangService:
    """
    Service layer cho Khách hàng
    Cung cấp các thao tác CRUD và business logic
    """
    
    def __init__(self, session: Session = None):
        """
        Khởi tạo service
        
        Args:
            session: Database session (optional, sẽ tự tạo nếu không có)
        """
        self.session = session
        self._external_session = session is not None
    
    def _get_session(self) -> Session:
        """Get session (tự tạo nếu chưa có)"""
        if self.session is None:
            return get_session()
        return self.session
    
    def _close_session(self, session: Session):
        """Close session nếu tự tạo"""
        if not self._external_session and session:
            session.close()
    
    def create(self, data: Dict[str, Any]) -> KhachHang:
        """
        Tạo khách hàng mới
        
        Args:
            data: Dictionary chứa thông tin khách hàng
                - ho_ten (str): Họ tên
                - loai_khach (str): 'ca_nhan' hoặc 'doanh_nghiep'
                - so_dien_thoai (str): Số điện thoại
                - email (str): Email
                - dia_chi (str): Địa chỉ
                - ma_so_thue (str, optional): Mã số thuế
                - ngay_dang_ky (date, optional): Ngày đăng ký
        
        Returns:
            KhachHang: Khách hàng đã tạo
        
        Raises:
            ValueError: Nếu dữ liệu không hợp lệ hoặc email đã tồn tại
            Exception: Nếu có lỗi database
        """
        session = self._get_session()
        try:
            # Validate data
            self._validate_data(data)
            
            # Check duplicate email
            if data.get('email'):
                self._check_duplicate_email(data['email'])
            
            # Generate mã khách hàng
            data['ma_khach_hang'] = self._generate_ma_khach_hang()
            
            # Create object
            khach_hang = KhachHang(**data)
            
            # Add to database
            session.add(khach_hang)
            session.commit()
            session.refresh(khach_hang)
            
            return khach_hang
            
        except Exception as e:
            session.rollback()
            raise e
        finally:
            self._close_session(session)
    
    def get_by_id(self, ma_khach_hang: str) -> Optional[KhachHang]:
        """
        Lấy khách hàng theo mã
        
        Args:
            ma_khach_hang: Mã khách hàng
        
        Returns:
            KhachHang hoặc None nếu không tìm thấy
        """
        session = self._get_session()
        try:
            return session.query(KhachHang).filter(
                KhachHang.ma_khach_hang == ma_khach_hang
            ).first()
        finally:
            self._close_session(session)
    
    def get_all(self, skip: int = 0, limit: int = 100, 
                trang_thai: str = None) -> List[KhachHang]:
        """
        Lấy tất cả khách hàng với pagination
        
        Args:
            skip: Số bản ghi bỏ qua
            limit: Số bản ghi tối đa
            trang_thai: Filter theo trạng thái (optional)
        
        Returns:
            List[KhachHang]: Danh sách khách hàng
        """
        session = self._get_session()
        try:
            query = session.query(KhachHang)
            
            # Filter by status if provided
            if trang_thai:
                query = query.filter(KhachHang.trang_thai == trang_thai)
            
            return query.offset(skip).limit(limit).all()
        finally:
            self._close_session(session)
    
    def search(self, keyword: str, skip: int = 0, 
               limit: int = 50) -> List[KhachHang]:
        """
        Tìm kiếm khách hàng theo từ khóa
        
        Args:
            keyword: Từ khóa tìm kiếm
            skip: Số bản ghi bỏ qua
            limit: Số bản ghi tối đa
        
        Returns:
            List[KhachHang]: Danh sách khách hàng tìm được
        """
        session = self._get_session()
        try:
            # Tạo search pattern
            search_pattern = f"%{keyword}%"
            
            # Tìm theo tên, SĐT, email
            results = session.query(KhachHang).filter(
                KhachHang.trang_thai != TrangThaiKHEnum.DA_XOA,
                or_(
                    KhachHang.ho_ten.ilike(search_pattern),
                    KhachHang.so_dien_thoai.ilike(search_pattern),
                    KhachHang.email.ilike(search_pattern),
                    KhachHang.ma_so_thue.ilike(search_pattern)
                )
            ).offset(skip).limit(limit).all()
            
            return results
        finally:
            self._close_session(session)
    
    def update(self, ma_khach_hang: str, 
               data: Dict[str, Any]) -> Optional[KhachHang]:
        """
        Cập nhật thông tin khách hàng
        
        Args:
            ma_khach_hang: Mã khách hàng cần update
            data: Dictionary chứa thông tin cần update
        
        Returns:
            KhachHang: Khách hàng đã update hoặc None nếu không tìm thấy
        
        Raises:
            ValueError: Nếu dữ liệu không hợp lệ
        """
        session = self._get_session()
        try:
            # Find customer within the same session (not via get_by_id to avoid detached object)
            khach_hang = session.query(KhachHang).filter(
                KhachHang.ma_khach_hang == ma_khach_hang
            ).first()
            
            if not khach_hang:
                return None
            
            # Validate data (excluding ma_khach_hang and ngay_dang_ky)
            exclude_fields = ['ma_khach_hang', 'ngay_dang_ky']
            validate_data = {k: v for k, v in data.items() if k not in exclude_fields}
            if validate_data:
                self._validate_data(validate_data, is_update=True)
            
            # Check duplicate email
            if 'email' in data and data['email']:
                self._check_duplicate_email(data['email'], exclude_ma=ma_khach_hang)
            
            # Update fields
            for key, value in data.items():
                if hasattr(khach_hang, key) and key not in exclude_fields:
                    setattr(khach_hang, key, value)
            
            # Update timestamp
            khach_hang.ngay_cap_nhat = datetime.now()
            
            # Commit
            session.commit()
            session.refresh(khach_hang)
            
            return khach_hang
            
        except Exception as e:
            session.rollback()
            raise e
        finally:
            self._close_session(session)
    
    def delete(self, ma_khach_hang: str) -> bool:
        """
        Xóa khách hàng (soft delete)
        
        Args:
            ma_khach_hang: Mã khách hàng cần xóa
        
        Returns:
            bool: True nếu xóa thành công, False nếu không tìm thấy
        
        Raises:
            ValueError: Nếu khách hàng còn hợp đồng đang hoạt động
        """
        session = self._get_session()
        try:
            # Find customer within the same session (not via get_by_id to avoid detached object)
            khach_hang = session.query(KhachHang).filter(
                KhachHang.ma_khach_hang == ma_khach_hang
            ).first()
            
            if not khach_hang:
                return False
            
            # Check for active contracts
            active_contracts = session.query(HopDong).filter(
                HopDong.ma_khach_hang == ma_khach_hang,
                HopDong.trang_thai == 'hieu_luc'
            ).count()
            
            if active_contracts > 0:
                raise ValueError(
                    f"Không thể xóa khách hàng có {active_contracts} hợp đồng đang hoạt động"
                )
            
            # Soft delete
            khach_hang.trang_thai = TrangThaiKHEnum.DA_XOA
            khach_hang.ngay_cap_nhat = datetime.now()
            
            session.commit()
            return True
            
        except Exception as e:
            session.rollback()
            raise e
        finally:
            self._close_session(session)
    
    def get_by_status(self, trang_thai: str, skip: int = 0, 
                      limit: int = 100) -> List[KhachHang]:
        """
        Lấy khách hàng theo trạng thái
        
        Args:
            trang_thai: Trạng thái ('hoat_dong', 'tam_khoa', 'da_xoa')
            skip: Số bản ghi bỏ qua
            limit: Số bản ghi tối đa
        
        Returns:
            List[KhachHang]: Danh sách khách hàng
        """
        session = self._get_session()
        try:
            return session.query(KhachHang).filter(
                KhachHang.trang_thai == trang_thai
            ).offset(skip).limit(limit).all()
        finally:
            self._close_session(session)
    
    def get_history(self, ma_khach_hang: str) -> Dict[str, Any]:
        """
        Lấy lịch sử khách hàng (hợp đồng, thanh toán)
        
        Args:
            ma_khach_hang: Mã khách hàng
        
        Returns:
            Dictionary chứa:
                - thong_tin: Thông tin khách hàng
                - hop_dongs: Danh sách hợp đồng
                - tong_hop_dong: Tổng số hợp đồng
                - hop_dong_dang_hieu_luc: Số hợp đồng đang hiệu lực
                - tong_da_thanh_toan: Tổng số tiền đã thanh toán
                - tong_cong_no: Tổng công nợ
        """
        session = self._get_session()
        try:
            khach_hang = self.get_by_id(ma_khach_hang)
            if not khach_hang:
                return {}
            
            # Get contracts
            hop_dongs = session.query(HopDong).filter(
                HopDong.ma_khach_hang == ma_khach_hang
            ).all()
            
            # Calculate statistics
            tong_hop_dong = len(hop_dongs)
            hop_dong_hieu_luc = sum(1 for hd in hop_dongs if hd.trang_thai == 'hieu_luc')
            
            # Get payments
            tong_da_thanh_toan = 0
            tong_cong_no = 0
            
            for hop_dong in hop_dongs:
                thanh_toans = session.query(ThanhToan).filter(
                    ThanhToan.ma_hop_dong == hop_dong.ma_hop_dong
                ).all()
                
                for tt in thanh_toans:
                    if tt.trang_thai == 'da_thanh_toan':
                        tong_da_thanh_toan += tt.so_tien
                    else:
                        tong_cong_no += tt.so_tien
            
            return {
                'thong_tin': khach_hang,
                'hop_dongs': hop_dongs,
                'tong_hop_dong': tong_hop_dong,
                'hop_dong_dang_hieu_luc': hop_dong_hieu_luc,
                'tong_da_thanh_toan': tong_da_thanh_toan,
                'tong_cong_no': tong_cong_no,
            }
            
        finally:
            self._close_session(session)
    
    def _check_duplicate_email(self, email: str, exclude_ma: str = None):
        """
        Check if email already exists
        
        Args:
            email: Email to check
            exclude_ma: Exclude this customer (for update)
        
        Raises:
            ValueError: If email already exists
        """
        session = self._get_session()
        try:
            query = session.query(KhachHang).filter(
                KhachHang.email == email,
                KhachHang.trang_thai != TrangThaiKHEnum.DA_XOA
            )
            
            if exclude_ma:
                query = query.filter(KhachHang.ma_khach_hang != exclude_ma)
            
            existing = query.first()
            if existing:
                raise ValueError(f"Email '{email}' đã được sử dụng bởi khách hàng {existing.ho_ten} ({existing.ma_khach_hang})")
        finally:
            self._close_session(session)
    
    def _validate_data(self, data: Dict[str, Any], is_update: bool = False):
        """
        Validate dữ liệu khách hàng
        
        Args:
            data: Dictionary chứa dữ liệu
            is_update: Nếu True thì bỏ qua validate required
        
        Raises:
            ValueError: Nếu dữ liệu không hợp lệ
        """
        errors = []
        
        # Validate required fields
        if not is_update:
            required_fields = ['ho_ten', 'so_dien_thoai', 'dia_chi']
            for field in required_fields:
                if field in data:
                    result = validate_required(data[field], field)
                    if not result:
                        errors.append(result.message)
        
        # Validate email
        if 'email' in data and data['email']:
            result = validate_email(data['email'])
            if not result:
                errors.append(result.message)
        
        # Validate phone
        if 'so_dien_thoai' in data:
            result = validate_phone(data['so_dien_thoai'])
            if not result:
                errors.append(result.message)
        
        if errors:
            raise ValueError("; ".join(errors))
    
    def _generate_ma_khach_hang(self) -> str:
        """
        Generate mã khách hàng tự động
        
        Format: KH + YYYYMMDD + XXX (3 số random)
        
        Returns:
            str: Mã khách hàng
        """
        from src.utils.helpers import generate_code
        return generate_code(prefix="KH", length=3, include_date=True, separator="")


# Convenience functions
def create_khach_hang(data: Dict[str, Any]) -> KhachHang:
    """Tạo khách hàng mới"""
    service = KhachHangService()
    return service.create(data)


def get_khach_hang(ma_khach_hang: str) -> Optional[KhachHang]:
    """Lấy khách hàng theo mã"""
    service = KhachHangService()
    return service.get_by_id(ma_khach_hang)


def get_all_khach_hang(skip: int = 0, limit: int = 100) -> List[KhachHang]:
    """Lấy tất cả khách hàng"""
    service = KhachHangService()
    return service.get_all(skip, limit)


def search_khach_hang(keyword: str, skip: int = 0, limit: int = 50) -> List[KhachHang]:
    """Tìm kiếm khách hàng"""
    service = KhachHangService()
    return service.search(keyword, skip, limit)


def update_khach_hang(ma_khach_hang: str, data: Dict[str, Any]) -> Optional[KhachHang]:
    """Cập nhật khách hàng"""
    service = KhachHangService()
    return service.update(ma_khach_hang, data)


def delete_khach_hang(ma_khach_hang: str) -> bool:
    """Xóa khách hàng"""
    service = KhachHangService()
    return service.delete(ma_khach_hang)


def export_khach_hang_to_excel(khach_hangs: list, file_path: str = None) -> str:
    """
    Export danh sách khách hàng ra Excel
    
    Args:
        khach_hangs: List of KhachHang objects
        file_path: Output file path (default: data/exports/khach_hang_YYYYMMDD.xlsx)
    
    Returns:
        str: Path to exported file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        import os
    except ImportError:
        raise ImportError("Please install openpyxl: pip install openpyxl")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Khách hàng"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "STT", "Mã KH", "Họ tên", "Loại", "Số điện thoại", 
        "Email", "Địa chỉ", "Mã số thuế", "Ngày đăng ký", "Trạng thái"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data
    for row, kh in enumerate(khach_hangs, 2):
        ws.cell(row=row, column=1, value=row-1).border = thin_border
        ws.cell(row=row, column=2, value=kh.ma_khach_hang).border = thin_border
        ws.cell(row=row, column=3, value=kh.ho_ten).border = thin_border
        ws.cell(row=row, column=4, value="Doanh nghiệp" if kh.loai_khach == LoaiKhachEnum.DOANH_NGHIEP else "Cá nhân").border = thin_border
        ws.cell(row=row, column=5, value=kh.so_dien_thoai).border = thin_border
        ws.cell(row=row, column=6, value=kh.email or "-").border = thin_border
        ws.cell(row=row, column=7, value=kh.dia_chi).border = thin_border
        ws.cell(row=row, column=8, value=kh.ma_so_thue or "-").border = thin_border
        ws.cell(row=row, column=9, value=kh.ngay_dang_ky.strftime("%d/%m/%Y") if kh.ngay_dang_ky else "-").border = thin_border
        
        trang_thai_cell = ws.cell(row=row, column=10, value=str(kh.trang_thai))
        trang_thai_cell.border = thin_border
        
        # Apply alignment to all cells in row
        for col in range(1, 11):
            ws.cell(row=row, column=col).alignment = cell_alignment
    
    # Auto-adjust column widths
    column_widths = [5, 15, 30, 15, 15, 25, 40, 15, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save file
    if file_path is None:
        from src.utils.helpers import get_data_dir
        export_dir = get_data_dir() / 'exports'
        os.makedirs(export_dir, exist_ok=True)
        file_path = export_dir / f"khach_hang_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb.save(str(file_path))
    return str(file_path)


__all__ = [
    'KhachHangService',
    'create_khach_hang',
    'get_khach_hang',
    'get_all_khach_hang',
    'search_khach_hang',
    'update_khach_hang',
    'delete_khach_hang',
    'export_khach_hang_to_excel',
]
